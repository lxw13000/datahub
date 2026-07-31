#!/usr/bin/env python3
"""将提现用户钻石分析接口的扁平 JSON 导出为 Excel 分析报告。"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


BUSINESS_TYPE_NAMES = {
    1: "赠送豪华礼物",
    2: "收到豪华礼物",
    3: "代理豪华礼物分账",
    4: "子代理豪华礼物分账",
    5: "房主豪华礼物分账",
    11: "赠送幸运礼物",
    12: "收到幸运礼物",
    13: "代理幸运礼物分账",
    14: "子代理幸运礼物分账",
    16: "幸运礼物中奖",
    17: "幸运礼物jackpot派奖",
    21: "游戏消费",
    22: "游戏反奖",
    23: "房主游戏分账",
    24: "代理游戏分账",
    25: "子代理游戏分账",
    26: "游戏jackpot派奖",
    31: "金币转账扣除",
    32: "转账到金币",
    33: "钻石兑换金币",
    34: "钻石兑换金币",
    35: "钻石兑换交易金币",
    36: "钻石兑换交易金币",
    37: "转账到交易金币",
    41: "购买装扮道具",
    42: "购买VIP",
    43: "直播间发红包",
    44: "直播间收红包",
    45: "直播间红包退回",
    51: "金币充值",
    52: "交易金币充值",
    53: "提现扣除",
    54: "提现退款",
    55: "提现订单奖励",
    61: "主播榜奖励",
    62: "每日签到奖励",
    63: "每日签到奖励",
    64: "累计签到奖励",
    65: "累计签到奖励",
    66: "日常观看直播奖励",
    67: "日常观看直播奖励",
    68: "日常关注主播奖励",
    69: "日常关注主播奖励",
    70: "日常充值奖励",
    71: "日常充值奖励",
    72: "日常榜单奖励",
    73: "日常榜单奖励",
    74: "直播时长奖励",
    75: "幸运礼物收益奖励",
    76: "日常幸运礼物消耗奖励",
    77: "日常幸运礼物消耗奖励",
    78: "日常游戏消耗奖励",
    79: "日常游戏消耗奖励",
    80: "日常发布动态奖励",
    81: "日常发布动态奖励",
    82: "新代理邀请奖励",
    83: "新代理邀请奖励",
    84: "新代理累计流水奖励",
    85: "新代理累计流水奖励",
    86: "邀请抽奖",
    87: "邀请抽奖",
    88: "vip日奖励",
    89: "日常上麦奖励",
    90: "日常上麦奖励",
    91: "优质主播时长奖励",
    92: "邀请榜单奖励",
    93: "代理超级主播收益奖励",
    94: "代理超级主播幸运礼物收益奖励",
    95: "超级主播收益奖励",
    96: "超级主播幸运礼物收益奖励",
    97: "幸运礼物消费榜单奖励",
    98: "幸运礼物收入榜单奖励",
    99: "游戏消费榜单奖励",
    100: "游戏收入榜单奖励",
    101: "代理收入榜单奖励",
    102: "世界杯竞猜投注",
    103: "世界杯竞猜奖励",
    104: "房间奖池奖励",
    105: "购买VIP套餐补签卡",
    106: "购买VIP套餐",
    107: "VIP签到奖励",
    108: "VIP任务奖励",
    109: "优质主播奖励",
    110: "PK榜单奖励",
    501: "运营调整",
    502: "运营调整",
    503: "运营调整",
    504: "运营调整",
    505: "运营调整",
    506: "运营调整",
}

STATUS_NAMES = {1: "加余额", -1: "减余额"}
TIME_FORMAT = "%Y-%m-%d %H:%M:%S"
INTEGER_PATTERN = re.compile(r"[+-]?\d+")

TITLE_FILL = PatternFill("solid", fgColor="1F4E78")
SECTION_FILL = PatternFill("solid", fgColor="D9EAF7")
HEADER_FILL = PatternFill("solid", fgColor="5B9BD5")
LIGHT_FILL = PatternFill("solid", fgColor="EAF2F8")
WHITE_FONT = Font(color="FFFFFF", bold=True)
HEADER_FONT = Font(color="FFFFFF", bold=True)
LABEL_FONT = Font(color="1F1F1F", bold=True)
BODY_FONT = Font(color="1F1F1F")
NEGATIVE_FONT = Font(color="C00000")
POSITIVE_FONT = Font(color="008000")
LIGHT_SIDE = Side(style="thin", color="D9E2F3")
BOTTOM_BORDER = Border(bottom=LIGHT_SIDE)
TOKEN_FORMAT = "#,##0"
USER_ID_FORMAT = "0"
PERCENT_FORMAT = "0.00%"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="将 /walletDiamond/withdrawAnalysis 返回的 JSON 导出为 Excel 分析报告。"
    )
    parser.add_argument("--input", required=True, type=Path, help="接口结果 JSON 文件")
    parser.add_argument("--output", required=True, type=Path, help="输出 .xlsx 文件")
    parser.add_argument(
        "--user-ids",
        type=Path,
        help="可选的提现用户 ID CSV，用于保留没有钻石流水的用户",
    )
    parser.add_argument("--withdraw-start", required=True, help="提现开始时间")
    parser.add_argument("--withdraw-end", required=True, help="提现结束时间")
    parser.add_argument("--statistics-start", required=True, help="钻石统计开始时间")
    parser.add_argument("--statistics-end", required=True, help="钻石统计结束时间")
    parser.add_argument(
        "--exclude-business-types",
        nargs="*",
        default=[],
        metavar="TYPE",
        help="可选；空格或逗号分隔的业务类型编号，例如 74 503 或 74,503",
    )
    return parser.parse_args()


def parse_integer(value: Any, field: str, location: str, *, minimum: int | None = None) -> int:
    """只接受真正的整数或整数字符串，绝不经由 float 转换。"""
    if type(value) is int:
        result = value
    elif isinstance(value, str) and INTEGER_PATTERN.fullmatch(value.strip()):
        result = int(value.strip())
    else:
        raise ValueError(f"{location} 的 {field} 必须是整数，实际值为 {value!r}")
    if minimum is not None and result < minimum:
        raise ValueError(f"{location} 的 {field} 不能小于 {minimum}，实际值为 {result}")
    return result


def reject_json_float(value: str) -> None:
    raise ValueError(f"JSON 中出现小数 {value}；tokens 和各编号字段必须使用整数")


def load_records(path: Path) -> list[dict[str, Any]]:
    if not path.is_file():
        raise ValueError(f"输入 JSON 不存在或不是文件：{path}")
    try:
        with path.open("r", encoding="utf-8-sig") as source:
            root = json.load(source, parse_int=int, parse_float=reject_json_float)
    except (OSError, json.JSONDecodeError, UnicodeDecodeError, ValueError) as error:
        raise ValueError(f"读取 JSON 失败：{error}") from error

    if isinstance(root, list):
        records = root
    elif isinstance(root, dict):
        if root.get("success") is False:
            raise ValueError("接口结果 success=false，不能生成报告")
        if "code" in root and parse_integer(root["code"], "code", "JSON 根节点") != 200:
            raise ValueError(f"接口结果 code={root['code']}，预期为 200")
        records = root.get("data")
        if not isinstance(records, list):
            raise ValueError("对象根节点必须包含数组类型的 data 字段")
    else:
        raise ValueError("JSON 根节点必须是数组，或包含 data 数组的对象")

    for index, record in enumerate(records, start=1):
        if not isinstance(record, dict):
            raise ValueError(f"第 {index} 条记录不是 JSON 对象")
    return records


def load_user_ids(path: Path | None) -> set[int]:
    if path is None:
        return set()
    if not path.is_file():
        raise ValueError(f"用户 ID 文件不存在或不是文件：{path}")

    user_ids: set[int] = set()
    header_checked = False
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as source:
            reader = csv.reader(source)
            for row_number, row in enumerate(reader, start=1):
                if not row or all(not cell.strip() for cell in row):
                    continue
                raw_value = row[0].strip()
                normalized = raw_value.lower().replace("_", "").replace("-", "")
                if not header_checked:
                    header_checked = True
                    if normalized in {"userid", "uid", "用户id", "用户编号"}:
                        continue
                user_ids.add(
                    parse_integer(
                        raw_value,
                        "userId",
                        f"用户 ID 文件第 {row_number} 行",
                        minimum=1,
                    )
                )
    except (OSError, UnicodeDecodeError, csv.Error) as error:
        raise ValueError(f"读取用户 ID 文件失败：{error}") from error
    return user_ids


def parse_excluded_business_types(values: Iterable[str]) -> set[int]:
    excluded: set[int] = set()
    for argument in values:
        for item in argument.split(","):
            item = item.strip()
            if item:
                excluded.add(
                    parse_integer(item, "businessType", "--exclude-business-types", minimum=0)
                )
    return excluded


def validate_times(args: argparse.Namespace) -> None:
    parsed: dict[str, datetime] = {}
    for field in ("withdraw_start", "withdraw_end", "statistics_start", "statistics_end"):
        value = getattr(args, field)
        try:
            parsed[field] = datetime.strptime(value, TIME_FORMAT)
        except ValueError as error:
            option = "--" + field.replace("_", "-")
            raise ValueError(f"{option} 格式错误，请使用 yyyy-MM-dd HH:mm:ss") from error
    if parsed["withdraw_end"] < parsed["withdraw_start"]:
        raise ValueError("提现结束时间不能早于提现开始时间")
    if parsed["statistics_end"] < parsed["statistics_start"]:
        raise ValueError("统计结束时间不能早于统计开始时间")


def aggregate_records(
    records: list[dict[str, Any]],
    extra_user_ids: set[int],
    excluded_business_types: set[int],
) -> tuple[dict[tuple[int, int, int], int], list[int]]:
    aggregated: dict[tuple[int, int, int], int] = defaultdict(int)
    all_user_ids = set(extra_user_ids)

    for index, record in enumerate(records, start=1):
        location = f"第 {index} 条记录"
        missing = [
            field
            for field in ("userId", "status", "businessType", "tokens")
            if field not in record
        ]
        if missing:
            raise ValueError(f"{location} 缺少字段：{', '.join(missing)}")
        user_id = parse_integer(record["userId"], "userId", location, minimum=1)
        status = parse_integer(record["status"], "status", location)
        business_type = parse_integer(
            record["businessType"], "businessType", location, minimum=0
        )
        tokens = parse_integer(record["tokens"], "tokens", location, minimum=0)
        if status not in STATUS_NAMES:
            raise ValueError(f"{location} 的 status 必须是 1 或 -1，实际值为 {status}")

        # 即使该记录随后被本地排除，仍保留其用户，避免丢失提现用户。
        all_user_ids.add(user_id)
        if business_type not in excluded_business_types:
            aggregated[(user_id, status, business_type)] += tokens

    return dict(aggregated), sorted(all_user_ids)


def business_name(business_type: int) -> str:
    return BUSINESS_TYPE_NAMES.get(business_type, "未知业务类型")


def build_analysis(
    aggregated: dict[tuple[int, int, int], int], user_ids: list[int]
) -> dict[str, Any]:
    details: dict[int, dict[int, dict[int, int]]] = {
        user_id: {1: {}, -1: {}} for user_id in user_ids
    }
    for (user_id, status, business_type), tokens in aggregated.items():
        details[user_id][status][business_type] = tokens

    users: list[dict[str, Any]] = []
    for user_id in user_ids:
        income_details = details[user_id][1]
        expense_details = details[user_id][-1]
        income = sum(income_details.values())
        expense = sum(expense_details.values())
        users.append(
            {
                "user_id": user_id,
                "income": income,
                "expense": expense,
                "net": income - expense,
                "income_details": income_details,
                "expense_details": expense_details,
            }
        )

    totals = {
        "user_count": len(users),
        "income": sum(user["income"] for user in users),
        "expense": sum(user["expense"] for user in users),
        "no_income": sum(user["income"] == 0 for user in users),
        "no_expense": sum(user["expense"] == 0 for user in users),
        "positive": sum(user["net"] > 0 for user in users),
        "zero": sum(user["net"] == 0 for user in users),
        "negative": sum(user["net"] < 0 for user in users),
    }
    totals["net"] = totals["income"] - totals["expense"]

    business_rows: list[dict[str, Any]] = []
    for status in (1, -1):
        business_types = sorted(
            {
                business_type
                for _, row_status, business_type in aggregated
                if row_status == status
            }
        )
        direction_total = totals["income"] if status == 1 else totals["expense"]
        for business_type in business_types:
            rows = [
                (user_id, tokens)
                for (user_id, row_status, row_type), tokens in aggregated.items()
                if row_status == status and row_type == business_type
            ]
            token_total = sum(tokens for _, tokens in rows)
            business_rows.append(
                {
                    "status": status,
                    "direction": STATUS_NAMES[status],
                    "business_type": business_type,
                    "name": business_name(business_type),
                    "user_count": len({user_id for user_id, _ in rows}),
                    "tokens": token_total,
                    "share": token_total / direction_total if direction_total else 0,
                }
            )
    business_rows.sort(
        key=lambda row: (
            0 if row["status"] == 1 else 1,
            -row["tokens"],
            row["business_type"],
        )
    )

    income_top = sorted(
        (user for user in users if user["income"] > 0),
        key=lambda user: (-user["income"], user["user_id"]),
    )[:20]
    expense_top = sorted(
        (user for user in users if user["expense"] > 0),
        key=lambda user: (-user["expense"], user["user_id"]),
    )[:20]
    net_low = sorted(users, key=lambda user: (user["net"], user["user_id"]))[:20]
    return {
        "details": details,
        "users": users,
        "totals": totals,
        "business_rows": business_rows,
        "income_top": income_top,
        "expense_top": expense_top,
        "net_low": net_low,
    }


def style_header(ws, row: int, start_column: int, end_column: int) -> None:
    for cell in ws.iter_cols(
        min_col=start_column,
        max_col=end_column,
        min_row=row,
        max_row=row,
    ):
        current = cell[0]
        current.fill = HEADER_FILL
        current.font = HEADER_FONT
        current.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        current.border = BOTTOM_BORDER
    ws.row_dimensions[row].height = 30


def style_section(ws, row: int, title: str, end_column: int) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_column)
    cell = ws.cell(row, 1, title)
    cell.fill = SECTION_FILL
    cell.font = Font(color="1F4E78", bold=True, size=12)
    cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 26
    for column in range(1, end_column + 1):
        ws.cell(row, column).fill = SECTION_FILL


def style_net_cell(cell) -> None:
    cell.number_format = TOKEN_FORMAT
    if cell.value < 0:
        cell.font = NEGATIVE_FONT
    elif cell.value > 0:
        cell.font = POSITIVE_FONT


def format_data_sheet(
    ws,
    widths: list[float],
    numeric_columns: Iterable[int],
    user_id_columns: Iterable[int] = (1,),
) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{max(1, ws.max_row)}"
    ws.sheet_view.showGridLines = False
    style_header(ws, 1, 1, ws.max_column)
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = BOTTOM_BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=cell.column >= 7)
            cell.font = BODY_FONT
        for column in numeric_columns:
            row[column - 1].number_format = TOKEN_FORMAT
        for column in user_id_columns:
            row[column - 1].number_format = USER_ID_FORMAT
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def detail_text(values: dict[int, int]) -> str:
    return "; ".join(
        f"{business_type}（{business_name(business_type)}）: {tokens:,}"
        for business_type, tokens in sorted(values.items())
    )


def write_summary_sheet(
    ws,
    analysis: dict[str, Any],
    args: argparse.Namespace,
    excluded_business_types: set[int],
    raw_record_count: int,
) -> None:
    totals = analysis["totals"]
    users = analysis["users"]
    business_rows = analysis["business_rows"]
    income_business = [row for row in business_rows if row["status"] == 1]
    expense_business = [row for row in business_rows if row["status"] == -1]

    ws.title = "分析汇总"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A6"
    ws.merge_cells("A1:N1")
    ws["A1"] = "提现用户钻石来源、用途分析"
    ws["A1"].fill = TITLE_FILL
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 38
    for column in range(1, 15):
        ws.cell(1, column).fill = TITLE_FILL

    ws["A2"] = "提现用户范围"
    ws.merge_cells("B2:E2")
    ws["B2"] = f"{args.withdraw_start} ～ {args.withdraw_end}"
    ws["F2"] = "钻石统计范围"
    ws.merge_cells("G2:J2")
    ws["G2"] = f"{args.statistics_start} ～ {args.statistics_end}"
    ws["K2"] = "生成时间"
    ws.merge_cells("L2:N2")
    ws["L2"] = datetime.now().strftime(TIME_FORMAT)

    excluded_text = "无"
    if excluded_business_types:
        excluded_text = "、".join(
            f"{item}（{business_name(item)}）" for item in sorted(excluded_business_types)
        )
    ws["A3"] = "排除业务类型"
    ws.merge_cells("B3:J3")
    ws["B3"] = excluded_text
    ws["K3"] = "输入 JSON"
    ws.merge_cells("L3:N3")
    ws["L3"] = args.input.name

    ws["A4"] = "提现用户 ID 文件"
    ws.merge_cells("B4:J4")
    ws["B4"] = args.user_ids.name if args.user_ids else "未提供"
    ws["K4"] = "原始/汇总记录数"
    ws.merge_cells("L4:N4")
    ws["L4"] = f"{raw_record_count:,} / {len(analysis['details']) and sum(len(v[1]) + len(v[-1]) for v in analysis['details'].values()):,}"

    for row in range(2, 5):
        for column in (1, 6, 11):
            ws.cell(row, column).font = LABEL_FONT
            ws.cell(row, column).fill = LIGHT_FILL
        for column in range(1, 15):
            ws.cell(row, column).alignment = Alignment(vertical="center", wrap_text=True)
            ws.cell(row, column).border = BOTTOM_BORDER
        ws.row_dimensions[row].height = 25

    first_headers = ["提现用户数", "加余额总数", "减余额总数", "净额", "无加余额用户数"]
    first_values = [
        totals["user_count"],
        totals["income"],
        totals["expense"],
        totals["net"],
        totals["no_income"],
    ]
    second_headers = [
        "无减余额用户数",
        "净额正用户数",
        "净额零用户数",
        "净额负用户数",
        "汇总明细数",
    ]
    second_values = [
        totals["no_expense"],
        totals["positive"],
        totals["zero"],
        totals["negative"],
        sum(len(value[1]) + len(value[-1]) for value in analysis["details"].values()),
    ]
    for column, (header, value) in enumerate(zip(first_headers, first_values), start=1):
        ws.cell(6, column, header)
        ws.cell(7, column, value)
    for column, (header, value) in enumerate(zip(second_headers, second_values), start=1):
        ws.cell(8, column, header)
        ws.cell(9, column, value)
    style_header(ws, 6, 1, 5)
    style_header(ws, 8, 1, 5)
    for row in (7, 9):
        for column in range(1, 6):
            cell = ws.cell(row, column)
            cell.number_format = TOKEN_FORMAT
            cell.font = Font(size=12, bold=True)
            cell.alignment = Alignment(horizontal="right")
            cell.border = BOTTOM_BORDER
    style_net_cell(ws["D7"])

    business_section = 11
    style_section(ws, business_section, "各方向业务类型构成", 14)
    business_header = business_section + 1
    income_headers = ["加余额业务类型", "涉及用户数", "Tokens总数", "方向内占比"]
    expense_headers = ["减余额业务类型", "涉及用户数", "Tokens总数", "方向内占比"]
    for column, header in enumerate(income_headers, start=1):
        ws.cell(business_header, column, header)
    for column, header in enumerate(expense_headers, start=6):
        ws.cell(business_header, column, header)
    style_header(ws, business_header, 1, 4)
    style_header(ws, business_header, 6, 9)

    business_data_start = business_header + 1
    for offset in range(max(len(income_business), len(expense_business), 1)):
        if offset < len(income_business):
            row = income_business[offset]
            values = [
                f"{row['business_type']}（{row['name']}）",
                row["user_count"],
                row["tokens"],
                row["share"],
            ]
            for column, value in enumerate(values, start=1):
                ws.cell(business_data_start + offset, column, value)
        if offset < len(expense_business):
            row = expense_business[offset]
            values = [
                f"{row['business_type']}（{row['name']}）",
                row["user_count"],
                row["tokens"],
                row["share"],
            ]
            for column, value in enumerate(values, start=6):
                ws.cell(business_data_start + offset, column, value)
        for column in list(range(1, 5)) + list(range(6, 10)):
            ws.cell(business_data_start + offset, column).border = BOTTOM_BORDER
    for row in range(business_data_start, business_data_start + max(len(income_business), 1)):
        ws.cell(row, 2).number_format = TOKEN_FORMAT
        ws.cell(row, 3).number_format = TOKEN_FORMAT
        ws.cell(row, 4).number_format = PERCENT_FORMAT
    for row in range(business_data_start, business_data_start + max(len(expense_business), 1)):
        ws.cell(row, 7).number_format = TOKEN_FORMAT
        ws.cell(row, 8).number_format = TOKEN_FORMAT
        ws.cell(row, 9).number_format = PERCENT_FORMAT

    top_section = business_data_start + max(len(income_business), len(expense_business), 1) + 1
    style_section(ws, top_section, "用户排名（最多 20 名）", 14)
    top_header = top_section + 1
    table_headers = [
        (1, ["加余额排名", "用户ID", "加余额总数", "净额"]),
        (6, ["减余额排名", "用户ID", "减余额总数", "净额"]),
        (11, ["净额最低排名", "用户ID", "净额", "加/减余额"]),
    ]
    for start_column, headers in table_headers:
        for column, header in enumerate(headers, start=start_column):
            ws.cell(top_header, column, header)
        style_header(ws, top_header, start_column, start_column + 3)

    top_start = top_header + 1
    for index in range(20):
        row_number = top_start + index
        if index < len(analysis["income_top"]):
            user = analysis["income_top"][index]
            values = [index + 1, user["user_id"], user["income"], user["net"]]
            for column, value in enumerate(values, start=1):
                ws.cell(row_number, column, value)
            ws.cell(row_number, 2).number_format = USER_ID_FORMAT
            ws.cell(row_number, 3).number_format = TOKEN_FORMAT
            style_net_cell(ws.cell(row_number, 4))
        if index < len(analysis["expense_top"]):
            user = analysis["expense_top"][index]
            values = [index + 1, user["user_id"], user["expense"], user["net"]]
            for column, value in enumerate(values, start=6):
                ws.cell(row_number, column, value)
            ws.cell(row_number, 7).number_format = USER_ID_FORMAT
            ws.cell(row_number, 8).number_format = TOKEN_FORMAT
            style_net_cell(ws.cell(row_number, 9))
        if index < len(analysis["net_low"]):
            user = analysis["net_low"][index]
            values = [
                index + 1,
                user["user_id"],
                user["net"],
                f"{user['income']:,} / {user['expense']:,}",
            ]
            for column, value in enumerate(values, start=11):
                ws.cell(row_number, column, value)
            ws.cell(row_number, 12).number_format = USER_ID_FORMAT
            style_net_cell(ws.cell(row_number, 13))
        for column in list(range(1, 5)) + list(range(6, 10)) + list(range(11, 15)):
            ws.cell(row_number, column).border = BOTTOM_BORDER

    max_section = top_start + 21
    style_section(ws, max_section, "最大用户占比及剔除影响", 14)
    max_header = max_section + 1
    max_headers = [
        "统计口径",
        "最大用户ID",
        "该方向数量",
        "方向总量",
        "最大用户占比",
        "剔除后加余额",
        "剔除后减余额",
        "剔除后净额",
        "说明",
    ]
    for column, header in enumerate(max_headers, start=1):
        ws.cell(max_header, column, header)
    style_header(ws, max_header, 1, len(max_headers))

    for offset, (direction, amount_key, total_key) in enumerate(
        (("加余额", "income", "income"), ("减余额", "expense", "expense")),
        start=1,
    ):
        row_number = max_header + offset
        direction_total = totals[total_key]
        candidates = [user for user in users if user[amount_key] > 0]
        selected = (
            sorted(candidates, key=lambda user: (-user[amount_key], user["user_id"]))[0]
            if candidates
            else None
        )
        if selected:
            remaining_income = totals["income"] - selected["income"]
            remaining_expense = totals["expense"] - selected["expense"]
            values = [
                f"剔除{direction}最大用户",
                selected["user_id"],
                selected[amount_key],
                direction_total,
                selected[amount_key] / direction_total if direction_total else 0,
                remaining_income,
                remaining_expense,
                remaining_income - remaining_expense,
                "从加、减两侧同时剔除该用户",
            ]
        else:
            values = [
                f"剔除{direction}最大用户",
                None,
                0,
                0,
                0,
                totals["income"],
                totals["expense"],
                totals["net"],
                f"{direction}方向无流水",
            ]
        for column, value in enumerate(values, start=1):
            ws.cell(row_number, column, value)
            ws.cell(row_number, column).border = BOTTOM_BORDER
        ws.cell(row_number, 2).number_format = USER_ID_FORMAT
        for column in (3, 4, 6, 7):
            ws.cell(row_number, column).number_format = TOKEN_FORMAT
        ws.cell(row_number, 5).number_format = PERCENT_FORMAT
        style_net_cell(ws.cell(row_number, 8))

    for column, width in {
        "A": 24,
        "B": 16,
        "C": 18,
        "D": 16,
        "E": 18,
        "F": 24,
        "G": 16,
        "H": 18,
        "I": 16,
        "J": 3,
        "K": 18,
        "L": 16,
        "M": 18,
        "N": 24,
    }.items():
        ws.column_dimensions[column].width = width
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0


def write_user_sheet(ws, analysis: dict[str, Any]) -> None:
    headers = [
        "用户ID",
        "加余额总数",
        "减余额总数",
        "净额（加-减）",
        "加余额业务数",
        "减余额业务数",
        "加余额业务明细",
        "减余额业务明细",
    ]
    ws.append(headers)
    for user in analysis["users"]:
        ws.append(
            [
                user["user_id"],
                user["income"],
                user["expense"],
                user["net"],
                len(user["income_details"]),
                len(user["expense_details"]),
                detail_text(user["income_details"]),
                detail_text(user["expense_details"]),
            ]
        )
    format_data_sheet(ws, [14, 18, 18, 18, 15, 15, 72, 72], (2, 3, 4))
    for row in range(2, ws.max_row + 1):
        style_net_cell(ws.cell(row, 4))
    if ws.max_row >= 2:
        ws.conditional_formatting.add(
            f"D2:D{ws.max_row}",
            CellIsRule(operator="lessThan", formula=["0"], font=NEGATIVE_FONT),
        )
        ws.conditional_formatting.add(
            f"D2:D{ws.max_row}",
            CellIsRule(operator="greaterThan", formula=["0"], font=POSITIVE_FONT),
        )


def write_detail_sheet(
    ws,
    aggregated: dict[tuple[int, int, int], int],
    status: int,
) -> None:
    ws.append(["用户ID", "业务类型", "业务类型备注", "Tokens"])
    rows = sorted(
        (
            (user_id, business_type, business_name(business_type), tokens)
            for (user_id, row_status, business_type), tokens in aggregated.items()
            if row_status == status
        ),
        key=lambda item: (item[0], item[1]),
    )
    for row in rows:
        ws.append(row)
    format_data_sheet(ws, [14, 12, 30, 20], (4,))


def write_business_sheet(ws, analysis: dict[str, Any]) -> None:
    ws.append(["方向", "业务类型", "业务类型备注", "涉及用户数", "Tokens总数", "方向内占比"])
    for row in analysis["business_rows"]:
        ws.append(
            [
                row["direction"],
                row["business_type"],
                row["name"],
                row["user_count"],
                row["tokens"],
                row["share"],
            ]
        )
    format_data_sheet(ws, [12, 12, 32, 16, 22, 16], (4, 5), user_id_columns=())
    for row in range(2, ws.max_row + 1):
        ws.cell(row, 6).number_format = PERCENT_FORMAT


def create_workbook(
    analysis: dict[str, Any],
    aggregated: dict[tuple[int, int, int], int],
    args: argparse.Namespace,
    excluded_business_types: set[int],
    raw_record_count: int,
) -> Workbook:
    workbook = Workbook()
    summary = workbook.active
    write_summary_sheet(
        summary, analysis, args, excluded_business_types, raw_record_count
    )
    write_user_sheet(workbook.create_sheet("用户汇总"), analysis)
    write_detail_sheet(workbook.create_sheet("加余额明细"), aggregated, 1)
    write_detail_sheet(workbook.create_sheet("减余额明细"), aggregated, -1)
    write_business_sheet(workbook.create_sheet("业务类型汇总"), analysis)

    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True
    workbook.calculation.calcMode = "auto"
    workbook.properties.title = "提现用户钻石来源、用途分析"
    workbook.properties.subject = "由 withdraw-diamond-analysis 本地工具生成"
    workbook.active = 0
    return workbook


def save_workbook(workbook: Workbook, output: Path) -> None:
    if output.suffix.lower() != ".xlsx":
        raise ValueError("--output 必须使用 .xlsx 扩展名")
    output.parent.mkdir(parents=True, exist_ok=True)
    temporary = output.with_name(f".{output.name}.tmp.xlsx")
    try:
        workbook.save(temporary)
        os.replace(temporary, output)
    finally:
        if temporary.exists():
            temporary.unlink()


def main() -> int:
    args = parse_args()
    try:
        validate_times(args)
        if args.output.resolve() == args.input.resolve():
            raise ValueError("--input 与 --output 不能指向同一文件")
        excluded = parse_excluded_business_types(args.exclude_business_types)
        records = load_records(args.input)
        user_ids = load_user_ids(args.user_ids)
        aggregated, all_user_ids = aggregate_records(records, user_ids, excluded)
        analysis = build_analysis(aggregated, all_user_ids)
        workbook = create_workbook(
            analysis, aggregated, args, excluded, len(records)
        )
        save_workbook(workbook, args.output)
    except ValueError as error:
        print(f"错误：{error}", file=sys.stderr)
        return 2
    except OSError as error:
        print(f"文件操作失败：{error}", file=sys.stderr)
        return 3

    totals = analysis["totals"]
    print(f"报告已生成：{args.output.resolve()}")
    print(
        "关键汇总："
        f"提现用户数={totals['user_count']:,}，"
        f"加余额={totals['income']:,}，"
        f"减余额={totals['expense']:,}，"
        f"净额={totals['net']:,}"
    )
    print(
        f"输入记录={len(records):,}，"
        f"合并并排除后记录={len(aggregated):,}，"
        f"排除业务类型={','.join(map(str, sorted(excluded))) or '无'}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
